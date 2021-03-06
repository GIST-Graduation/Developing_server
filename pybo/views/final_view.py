from django.shortcuts import render, redirect
from django.contrib import messages
import logging
import sys, io
import os

# 액셀, csv를 다루는 모듈 --> pip install openpyxl
import openpyxl, csv

# 파이썬 글자색 적용 모듈 --> pip install termcolor
from termcolor import colored

os.getcwd()


def graduation_check(request):
    major_choice = {
        ("0", "physics_core"),
        ("1", "chemical_core"),
        ("2", "biology_core"),
        ("3", "eecs_core"),
        ("4", "mechanics_core"),
        ("5", "material_core"),
        ("6", "environment_core"),
    }
    year_choice = {
        ("16", "16학번"),
        ("17", "17학번"),
        ("18", "18학번"),
        ("19", "19학번"),
        ("20", "20학번"),
        ("21", "21학번"),
    }

    # 학번
    # 어떻게 하는지 잘 모르지만 일단 적어둠...
    my_major = 3
    incoming_year = 9

    my_classified_courses_credit = {
        **dict.fromkeys(
            ["core_english1", "core_english2", "core_writing",
             "HUS", "PPE", "humanity",
             "software",
             "core_math1", "core_math2",
             "core_science_ch", "core_experiment_ch",
             "core_science_ph", "core_experiment_ph",
             "core_science_bs", "core_experiment_bs",
             "core_science_cp",
             "freshman_seminar",
             "major",
             "research",
             "uc_core", "uc_optional1", "uc_optional2", "others",
             "music", "exercise", "colloquium",
             "nonclassified_courses", "F or U"], [0, 0])}

    # 수강한 강좌의 정보를 담는 딕셔너리
    # (강의코드, 강의학점, 강의명)
    my_classified_courses = {
        **dict.fromkeys(
            ["core_english1", "core_english2", "core_writing",
             "HUS", "PPE", "humanity",
             "software",
             "core_math1", "core_math2",
             "core_science_ch", "core_experiment_ch",
             "core_science_ph", "core_experiment_ph",
             "core_science_bs", "core_experiment_bs",
             "core_science_cp",
             "freshman_seminar",
             "major",
             "research",
             "uc_core", "uc_optional1", "uc_optional2", "others",
             "music", "exercise", "colloquium",
             "nonclassified_courses", "F or U"], [])}

    # 카테고리별 이수 여부
    # 이수 o = True
    # 이수 x = False
    category_complete_status = {
        **dict.fromkeys(
            ["core_english1", "core_english2", "core_writing",
             "HUS", "PPE", "humanity",
             "software",
             "core_math1", "core_math2",
             "core_science_ch", "core_experiment_ch",
             "core_science_ph", "core_experiment_ph",
             "core_science_bs", "core_experiment_bs",
             "core_science_cp",
             "freshman_seminar",
             "major",
             "research",
             "uc_core", "uc_optional1", "uc_optional2", "others",
             "music", "exercise", "colloquium"], False)}

    # 대분류별 이수 여부
    # 이수 o = True
    # 이수 x = False
    # 왜 True로 시작?
    # => 대분류에 속한 category 중에서 하나라도 False면 True를 False로 바꿔준다
    large_category_complete_status = {
        **dict.fromkeys(
            ["언어의 기초", "교양", "소프트웨어", "기초수학", "기초과학",
             "신입생 세미나", "전공", "연구", "대학 공통", "자유학점",
             "예체능", "콜로퀴움", "F or U", "분류 안 됨"], True)}

    # 카테고리별 최소 요구, 최대 인정 학점
    classified_courses_credit = {
        "core_english1": [2, 2],
        "core_english2": [2, 2],
        "core_writing": [3, 3],
        "HUS": [6, 6],
        "PPE": [6, 6],
        "humanity_core": [12, 24],
        "core_math1": [3, 3],
        "core_math2": [3, 3],
        "core_science_ch": [3, 3],
        "core_experiment_ch": [1, 1],
        "core_science_ph": [3, 3],
        "core_experiment_ph": [1, 1],
        "core_science_bs": [3, 3],
        "core_experiment_bs": [1, 1],
        "core_science_cp": [3, 3],
        "software": [2, 2],
        "freshman_seminar": [1, 1],
        # 신소재 [30, 42] 나머지 [36, 42]
        # 오류나면 수정좀
        "major": [36, 42] if (my_major) == "5" else [30, 42],
        "research": [6, 6],
        "uc_core": [1, 1],
        "uc_optional1": [0, 2],
        "uc_optional2": [0, 1],
        "others": '-',
        # 19학번까지는 예체능 각 4학기, 20학번부터는 각 2학기
        # 이것도 수정좀
        "music": [4, 4] if (incoming_year < 20) else [2, 2],
        "exercise": [4, 4] if (incoming_year < 20) else [2, 2],
        "colloquium": [2, 2],
        "nonclassified_courses": '-'
    }

    # 강좌 이름 딕셔너리
    courses_name = {
        "core_english1": "영어 1",
        "core_english2": "영어 2",
        "core_writing": "글쓰기",
        "HUS": "HUS",
        "PPE": "PPE",
        "humanity": "교양 최소 12학점 최대 24학점",
        "software": "소프트웨어(소기코)",
        "core_math1": "Core Mathematics 1",
        "core_math2": "Core Mathematics 2",
        "core_science_ch": "화학1(일반, 고급)",
        "core_science_ph": "물리학1(일반, 고급)",
        "core_science_bs": "생물학(일반, 인간, 고급)",
        "core_science_cp": "컴퓨터 프로그래밍",
        "core_experiment_ch": "일반화학 실험1",
        "core_experiment_ph": "일반물리학 실험1",
        "core_experiment_bs": "일반생물학 실험",
        "freshman_seminar": "신입생 세미나",
        "major": "전공(전필+전선)",
        "research": "학사논문연구",
        "uc_core": "과기경",
        "uc_optional1": "대학공통 선택(창의함양, 정신건강 관리)",
        "uc_optional2": "대학공통 선택(사회봉사, 해외봉사)",  # 최대 1학점 인정
        "others": "자유선택 - 언어선택, 소프트웨어, 기초과학선택, 기초전공, 타 전공, 부 전공",
        "music": "예능실기",
        "exercise": "체육실기",
        "colloquium": "GIST대학 콜로퀴움",
        "nonclassified_courses": "분류되지 않은 과목",
        "F or U": "F(U)학점을 맞은 과목"
    }

    major = [["physics_core", "physics_elective"],
             ["chemical_core", "chemical_elective"],
             ["biology_core", "biology_elective"],
             ["eecs_core", "eecs_elective"],
             ["mechanics_core", "mechanics_elective"],
             ["material_core", "material_elective"],
             ["environment_core", "environment_elective"]]

    my_classified_course = ["core_english1", "core_english2", "core_writing",
                            "HUS", "PPE", "humanity",
                            "software",
                            "core_math1", "core_math2",
                            "core_science_ch", "core_experiment_ch",
                            "core_science_ph", "core_experiment_ph"
                                               "core_science_bs", "core_experiment_bs",
                            "core_science_cp",
                            "freshman_seminar",
                            "major",
                            "research",
                            "uc_core", "uc_optional1", "uc_optional2", "others",
                            "music", "exercise", "colloquium", "nonclassified_courses", "F or U"]

    large_category = {
        "언어의 기초": ["core_english1", "core_english2", "core_writing"],
        "교양": ["HUS", "PPE", "humanity_core", "humanity_optional"],
        "소프트웨어": ["software"],
        "기초수학": ["core_math1", "core_math2"],
        "기초과학": ["core_science_ch", "core_experiment_ch",
                 "core_science_ph", "core_experiment_ph",
                 "core_science_bs", "core_experiment_bs",
                 "core_science_cp"],
        "신입생 세미나": ["freshman_seminar"],
        "전공": ["major"],
        "연구": ["research"],
        "대학 공통": ["uc_core", "uc_optional1", "uc_optional2"],
        "자유학점": ["others"],
        "예체능": ["music", "exercise"],
        "콜로퀴움": ["colloquium"],
        "분류 안 됨": ["nonclassified_courses"],
        "F or U": ["F or U"],
    }

    classified_courses = {
        'core_english1': ['GS1601', 'GS1603'],
        'core_english2': ['GS2652'],
        'core_writing': ['GS1511', 'GS1512', 'GS1513', 'GS1531', 'GS1532', 'GS1533'],
        'HUS': ['GS2501', 'GS2503', 'GS2505', 'GS2506', 'GS2507', 'GS2509', 'GS2510', 'GS2511', 'GS2512', 'GS2521',
                'GS2522', 'GS2523', 'GS2524', 'GS2525', 'GS2526', 'GS2544', 'GS2601', 'GS2602', 'GS2603', 'GS2604',
                'GS2611', 'GS2612', 'GS2613', 'GS2614', 'GS2615', 'GS2616', 'GS2618', 'GS2621', 'GS2622', 'GS2623',
                'GS2625', 'GS2626', 'GS2627', 'GS2628', 'GS2629', 'GS2630', 'GS2814', 'GS3501', 'GS3502', 'GS3504',
                'GS3601', 'GS3602', 'GS3603', 'GS3604', 'GS3621', 'GS3622', 'GS3623', 'GS3624', 'GS3625', 'GS3626',
                'GS3662', 'GS3801', 'GS3802', 'GS3803', 'GS3901'],
        'PPE': ['GS2620', 'GS2661', 'GS2701', 'GS2702', 'GS2703', 'GS2704', 'GS2705', 'GS2706', 'GS2707', 'GS2708',
                'GS2709', 'GS2724', 'GS2725', 'GS2726', 'GS2727', 'GS2728', 'GS2729', 'GS2730', 'GS2731', 'GS2732',
                'GS2733', 'GS2734', 'GS2735', 'GS2736', 'GS2742', 'GS2743', 'GS2747', 'GS2748', 'GS2750', 'GS2751',
                'GS2752', 'GS2761', 'GS2762', 'GS2763', 'GS2764', 'GS2765', 'GS2766', 'GS2781', 'GS2782', 'GS2783',
                'GS2784', 'GS2785', 'GS2786', 'GS2787', 'GS2788', 'GS2803', 'GS2812', 'GS2831', 'GS2832', 'GS2833',
                'GS2834', 'GS2835', 'GS3631', 'GS3632', 'GS3633', 'GS3661', 'GS3663', 'GS3721', 'GS3751', 'GS3752',
                'GS3753', 'GS3762', 'GS3763', 'GS3764', 'GS3861', 'GS4741', 'GS4761', 'GS4762'],
        'humanity': ['GS2541', 'GS2542', 'GS2543', 'GS2544', 'GS2791', 'GS2792', 'GS2793', 'GS2804', 'GS2808',
                     'GS2810', 'GS2815', 'GS2816', 'GS2817', 'GS2818', 'GS2819', 'GS2821', 'GS2822', 'GS2911',
                     'GS2912', 'GS2913', 'GS2931', 'GS2932', 'GS2933', 'GS3566'],
        'software': ['GS1490'],
        'core_math1': ['GS1001', 'GS1011'],
        'core_math2': ['GS2001', 'GS2002', 'GS2004', 'GS2013'],
        "core_science_ch": ['GS1201', 'GS1203'],
        "core_science_ph": ['GS1101', 'GS1103'],
        "core_science_bs": ['GS1301', 'GS1302', 'GS1303'],
        "core_experiment_ch": ['GS1211'],
        "core_experiment_ph": ['GS1111'],
        "core_experiment_bs": ['GS1311'],
        "core_science_cp": ['GS1401'],
        'freshman_seminar': ['GS1901', 'GS9301'],
        'physics_core': ['PS2101', 'PS2102', 'PS2103', 'PS3101', 'PS3103', 'PS3104', 'PS3105', 'PS3106', 'PS3107'],
        'physics_elective': ['PS2201', 'PS2202', 'PS3202', 'PS3203', 'PS3205', 'PS3206', 'PS4202', 'PS4203', 'PS4204',
                             'PS4205', 'PS4206', 'PS4207', 'PS4208', 'PS4209', 'PS4210', 'PS4211', 'PS4212', 'PS4213',
                             'PS4214', 'PS4215', 'PS4216'],
        'chemical_core': ['CH2101', 'CH2102', 'CH2103', 'CH2104', 'CH2105', 'CH3102', 'CH3103', 'CH3104', 'CH3105',
                          'CH3106', 'CH3107'],
        'chemical_elective': ['CH2106', 'CH2201', 'CH3202', 'CH3204', 'CH3205', 'CH3207', 'CH4205', 'CH4211', 'CH4212',
                              'CH4213', 'CH4215', 'CH4216', 'CH4218', 'CH4219', 'CH4220', 'CH4221', 'CH4222', 'CH4223'],
        'biology_core': ['BS2101', 'BS2102', 'BS3101', 'BS3105', 'BS3111', 'BS3112', 'BS3113'],
        'biology_elective': ['BS2201', 'BS3201', 'BS3202', 'BS3204', 'BS3205', 'BS4201', 'BS4202', 'BS4204', 'BS4205',
                             'BS4206', 'BS4207', 'BS4211', 'BS4212', 'BS4213', 'BS4214', 'BS4215', 'BS4216', 'BS4217',
                             'BS4218'],
        'eecs_core': ['EC3101', 'EC3102'],
        'eecs_elective': ['EC2105', 'EC2201', 'EC2202', 'EC2203', 'EC2204', 'EC2205', 'EC2206', 'EC3102', 'EC3202',
                          'EC3204', 'EC3206', 'EC3207', 'EC3208', 'EC3212', 'EC3213', 'EC3214', 'EC3215', 'EC3216',
                          'EC3217', 'EC3218', 'EC4202', 'EC4203', 'EC4204', 'EC4205', 'EC4206', 'EC4207', 'EC4208',
                          'EC4209', 'EC4210', 'EC4211', 'EC4212', 'EC4213', 'EC4214', 'EC4215', 'EC4216', 'EC4217',
                          'EC4218', 'EC4219', 'EC4301', 'EC4302'],
        'mechanics_core': ['MC2100', 'MC2101', 'MC2102', 'MC2103', 'MC3103', 'MC3105', 'MC3212', 'MC4101'],
        'mechanics_elective': ['MC3201', 'MC3202', 'MC3203', 'MC3204', 'MC3205', 'MC3206', 'MC3207', 'MC3208', 'MC3209',
                               'MC3210', 'MC3211', 'MC4202', 'MC4204', 'MC4205', 'MC4206', 'MC4208', 'MC4209', 'MC4210',
                               'MC4211', 'MC4212', 'MC4213', 'MC4214', 'MC4215', 'MC4216', 'MC4217', 'MC4218', 'MC4219',
                               'MC4221'],
        'material_core': ['MA2101', 'MA2102', 'MA2103', 'MA2104', 'MA3101', 'MA3102', 'MA3104', 'MA3105'],
        'material_elective': ['MA2201', 'MA2202', 'MA3201', 'MA3202', 'MA3203', 'MA3204', 'MA3205', 'MA3207', 'MA3208',
                              'MA3209', 'MA3210', 'MA3211', 'MA4201', 'MA4202', 'MA4203', 'MA4204', 'MA4205', 'MA4206',
                              'MA4207', 'MA4208', 'MA4209', 'MA4210', 'MA4211', 'MA4212', 'MA4213', 'MA4214', 'MA4215',
                              'MA4216', 'MA4217', 'MA4218', 'MA4219', 'MA4220', 'MA4221'],
        'environment_core': ['EV3101', 'EV3104', 'EV3105', 'EV3106', 'EV3111', 'EV4105', 'EV4106', 'EV4107'],
        'environment_elective': ['EV2208', 'EV2209', 'EV2210', 'EV2211', 'EV3205', 'EV3208', 'EV3213', 'EV3214',
                                 'EV3215',
                                 'EV3216', 'EV3217', 'EV3218', 'EV3219', 'EV3220', 'EV4201', 'EV4202', 'EV4203',
                                 'EV4204',
                                 'EV4205', 'EV4206', 'EV4209', 'EV4210', 'EV4211', 'EV4212', 'EV4213', 'EV4214',
                                 'EV4215',
                                 'EV4216', 'EV4217', 'EV4218', 'EV4221', 'EV4222', 'EV4223', 'EV4224', 'EV4225'],
        'research': ['9102', '9103', '9104'],
        'uc_core': ['UC0901'],
        "uc_optional1": ['UC0202', 'UC0301'],
        "uc_optional2": ['UC0201', 'UC0203'],
        'others': ['CT2501', 'CT2502', 'CT2503', 'CT2504', 'CT2505', 'CT2506', 'CT4101', 'CT41__', 'CT4201', 'CT4301',
                   'CT4302', 'CT4501', 'CT4502', 'CT4503', 'CT4504', 'CT4506', 'CT45__', 'ET2101', 'ET4102', 'ET4201',
                   'ET4302', 'ET4303', 'ET4304', 'ET4305', 'ET4306', 'ET4501', 'GS1102', 'GS1104', 'GS1112', 'GS1202',
                   'GS1204', 'GS1212', 'GS1321', 'GS1402', 'GS1431', 'GS1451', 'GS1471', 'GS1491', 'GS1605', 'GS1606',
                   'GS2006', 'GS2007', 'GS2102', 'GS2103', 'GS2104', 'GS2201', 'GS2202', 'GS2204', 'GS2206', 'GS2302',
                   'GS2303', 'GS2304', 'GS2311', 'GS2401', 'GS2402', 'GS2403', 'GS2406', 'GS2407', 'GS2408', 'GS2434',
                   'GS2435', 'GS2451', 'GS2471', 'GS2472', 'GS2473', 'GS2651', 'GS2653', 'GS2654', 'GS2655', 'GS2806',
                   'GS2809', 'GS2811', 'GS3001', 'GS3012', 'GS3015', 'GS3301', 'GS3311', 'GS3651', 'GS4002', 'GS4003',
                   'GS4004', 'GS4005', 'GS4006', 'GS4007', 'GS4008', 'GS4009', 'GS4010', 'GS4015', 'GS4016', 'GS4017',
                   'GS4018', 'GS4019', 'GS4301', 'IR2201', 'IR2202', 'IR3201', 'IR3202', 'IR3203', 'IR3204', 'IR4201',
                   'IR4202', 'IR4203', 'IR4204', 'IR4205', 'IR4206', 'IR4207', 'IR4208', 'IR4209', 'MD2101', 'MD4101',
                   'MD4102', 'MD4301', 'MD4302', 'MD4303', 'MD4501', 'MD4502', 'MD4601', 'MM2001', 'MM2002', 'MM2004',
                   'MM2006', 'MM2007', 'MM2011', 'MM3001', 'MM3012', 'MM3015', 'MM4002', 'MM4003', 'MM4004', 'MM4005',
                   'MM4006', 'MM4007', 'MM4008', 'MM4009', 'MM4010', 'MM4015', 'MM4016', 'MM4017', 'MM4018', 'MM4019',
                   ],
        'exercise': ['GS0101', 'GS0102', 'GS0103', 'GS0104', 'GS0105', 'GS0106', 'GS0107', 'GS0108', 'GS0109', 'GS0110',
                     'GS0111', 'GS0112', 'GS0113', 'GS0114'],
        'music': ['GS0201', 'GS0202', 'GS0203', 'GS0204', 'GS0205', 'GS0206', 'GS0207', 'GS0208', 'GS0209', 'GS0210',
                  'GS0211', 'GS0212'],
        'colloquium': ['GS9331', 'UC9331']
    }

    def get_my_courses(ws):
        index = 6
        result = []

        while True:
            # 코드, 강의명, 학점, 성적
            code = ws['B' + str(index)].value
            title = ws['D' + str(index)].value
            credit = ws['E' + str(index)].value
            grade = ws['F' + str(index)].value

            # 액셀 파일에서 B행 [학사]가 나오는 부분에서
            # 과목코드가 끝남 코드도 그렇게 설계
            if code == '[학사]':
                break

            if code == None or credit == None or title == None:
                index += 1
                continue

            # F학점은 따로 담아주고 result에는 반영 X
            if (grade == 'F' or grade == "U"):
                my_classified_courses["F"].append(
                    (code, int(credit), title))
                index += 1
                continue

            # result라는 배열에 (과목코드. 학점, 과목명) 튜플을 넣음
            result.append((code, int(credit), title))
            index += 1

        # lambda 함수를 이용해 학점 내림차순으로 정렬
        result.sort(key=lambda x: x[0], reverse=True)
        print("Done", flush=True)

        return result

    def classify_my_course(my_course_index, my_courses):
        # my_course = (code, credit, title)
        my_course = my_courses[my_course_index]

        # 글쓰기, 수학(다변수, 선대, 미방, 기초미방과 선대)
        for category in ["core_writing", "core_math2"]:

            if my_course[0] in classified_courses[category]:

                # 초과해서 들으면 자유 학점으로 처리
                if my_course[1] + my_classified_courses_credit[category][1] > classified_courses_credit[category][1]:
                    my_classified_courses_credit["others"][1] += my_course[1]
                    my_classified_courses["others"].append(my_course)
                # 초과 안하면 기초 필수 과목으로 분류
                else:
                    my_classified_courses_credit[1][category] += my_course[1]
                    my_classified_courses[category].append(my_course)

                return True

        # 영어12, 미적분학, 물화, 물화생 실험, 컴프, 신입생 세미나, 과기경, 대학공통 선택, 기타 등등
        for category in ["core_english1", "core_english2", "core_math1", "software",
                         "core_experiment_ch", "core_experiment_ph", "core_experiment_bs",
                         "core_science_cp", "core_science_ch", "core_science_ph",
                         "freshman_seminar", "uc_core", "uc_optional1", "uc_optional2", "others"]:

            # 컴프를 들었을 때 소기코 필수 최소 학점을 0으로
            if my_course[0] == "GS1401":
                classified_courses_credit["software"] = [0, 2]

            # 내 강좌의 과목코드가 (분류) 강좌에 있다면 딕셔너리 value의 학점 더하기 및 튜플 추가
            if my_course[0] in classified_courses[category]:
                my_classified_courses_credit[category][1] += my_course[1]
                my_classified_courses[category].append(my_course)

                return True

        # 생물은 따로 => 생물학, 인간생물학을 둘다 들을 경우 하나는 자유학점으로 넣어야 하기 때문
        if my_course[0] in classified_courses["core_science_bs"]:
            # 생물학, 인간생물학 둘다 들은 경우
            if my_classified_courses_credit["core_science_bs"][1] == 3:
                my_classified_courses_credit["others"][1] += my_course[1]
                my_classified_courses["others"].append(my_course)
            # 하나만 들은 경우(또는 둘 다 들었는데 아직 하나만 넣은 경우)
            else:
                my_classified_courses_credit["core_science_bs"][1] += my_course[1]
                my_classified_courses["core_science_bs"].append(my_course)

        # HUS, PPE, GSC
        for category in ["HUS", "PPE", "humanity"]:

            # GSC
            if category == "humanity":
                if my_course[0] in classified_courses[category]:
                    my_classified_courses_credit[category][1] += my_course[1]
                    my_classified_courses[category].append(my_course)

                    return True

            else:
                if my_course[0] in classified_courses[category]:

                    # HUS나 PPE 필수 학점을 넘었다면
                    if my_course[1] + my_classified_courses_credit[category][1] > \
                            classified_courses_credit[category][1]:

                        # 교양(12~24학점)에 넣는다
                        my_classified_courses_credit["humanity"][1] += my_course[1]
                        my_classified_courses["humanity"].append(
                            my_course)

                    # 안 넘었다면
                    else:
                        # HUS나 PPE에 넣는다
                        my_classified_courses_credit[category][1] += my_course[1]
                        my_classified_courses[category].append(my_course)

                    return True

        # 내 전공 분류
        # major = [[전필, 전선], [전필, 전선],....] ex) major[0] = ["physics_core", "physics_elective"]
        # classified_category = "내 전공_core", "내 전공_optional"
        # category = "전필", "전선"
        for classified_category, category in zip(major[int(my_major)], ["major_core", "major_elective"]):
            if my_course[0] in classified_courses[classified_category]:
                my_classified_courses_credit["major"][1] += my_course[1]
                my_classified_courses["major"].append(my_course)

                return True

        # 타 전공 분류
        for classified_category in [category for sublist in
                                    major[:int(my_major)] + major[int(my_major) + 1:] for category in
                                    sublist]:
            if my_course[0] in classified_courses[classified_category]:
                my_classified_courses_credit["others"][1] += my_course[1]
                my_classified_courses["others"].append(my_course)

                return True

        # 연구 학점
        for code in classified_courses["research"]:
            if my_course[0][2:] == code:
                my_classified_courses_credit["research"][1] += my_course[1]
                my_classified_courses["research"].append(my_course)

                return True

        # 예체능, 콜로퀴움
        for category in ["music", "exercise", "colloquium"]:
            if my_course[0] in classified_courses[category]:
                my_classified_courses_credit[category][1] += 1
                my_classified_courses[category].append(my_course)

                return True

        # 위에 과정을 모두 거쳐도 분류가 되지 않은 경우
        my_classified_courses_credit["nonclassified_courses"][1] += my_course[1]
        my_classified_courses["nonclassified_courses"].append(my_course)

        return True

    # 기초과학 수정
    # 기초과학 이수 조건이 충족 됐으면 과학 세 개와 컴프, 소기코 모두 True여야 함
    def modifying_basic_science():
        # 컴프를 들었을 경우
        # 셋 중 둘만 충족하면 됨
        if my_classified_courses_credit["core_science_cp"][1] == 3:

            category_complete_status["core_science_cp"] = True

            # 화학
            if my_classified_courses_credit["core_science_ch"][1] == 3:
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_science_ch"] = True
                    category_complete_status["core_experiment_ch"] = True

            # 물리
            if my_classified_courses_credit["core_science_ph"][1] == 3:
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_science_ph"] = True
                    category_complete_status["core_experiment_ph"] = True

            # 화학, 물리로 조건 충족한 경우
            if category_complete_status["core_science_ph"] == True and category_complete_status[
                "core_science_ch"] == True:
                # 생물학, 생실을 자유학점에 넣어준다
                my_classified_courses_credit["others"][1] += my_classified_courses_credit["core_science_bs"][1] + \
                                                             my_classified_courses_credit["core_experiment_bs"][1]
                my_classified_courses["others"].append(
                    my_classified_courses["core_science_bs"])
                my_classified_courses["others"].append(
                    my_classified_courses["core_experiment_bs"])

                # 최소, 최대 이수학점 0으로 만들어주고 이수 여부 true로
                classified_courses_credit["core_science_bs"] = [0, 0]
                classified_courses_credit["core_experiment_bs"] = [0, 0]
                category_complete_status["core_science_bs"] = True
                category_complete_status["core_experiment_bs"] = True

                # 생물학, 생실 삭제
                my_classified_courses_credit["core_science_bs"][1] = 0
                my_classified_courses_credit["core_experiment_bs"][1] = 0
                my_classified_courses["core_science_bs"] = []
                my_classified_courses["core_experiment_bs"] = []

                category_complete_status["software"] = True

                return True

            # 화학, 물리로는 기초과학 이수 조건을 충족하지 못한 경우
            else:
                # 생물
                if my_classified_courses_credit["core_science_bs"][1] == 3:
                    if my_classified_courses_credit["core_experiment_bs"][1] == 1:
                        category_complete_status["core_science_bs"] = True
                        category_complete_status["core_experiment_bs"] = True

                        # 화학 + 생물 조합
                        if category_complete_status["core_science_ch"] == True:
                            # 물리, 물실을 자유학점에 넣어준다
                            my_classified_courses_credit["others"][1] += \
                                my_classified_courses_credit["core_science_ph"][1] + \
                                my_classified_courses_credit["core_experiment_ph"][1]
                            my_classified_courses["others"].append(
                                my_classified_courses["core_science_ph"])
                            my_classified_courses["others"].append(
                                my_classified_courses["core_experiment_ph"])

                            # 물리, 물실 삭제
                            my_classified_courses_credit["core_science_ph"][1] = 0
                            my_classified_courses_credit["core_experiment_ph"][1] = 0
                            my_classified_courses["core_science_ph"] = []
                            my_classified_courses["core_experiment_ph"] = [
                            ]

                            # 최소, 최대 이수학점 0으로 만들어주고 이수 여부 true로
                            classified_courses_credit["core_science_ph"] = [
                                0, 0]
                            classified_courses_credit["core_experiment_ph"] = [
                                0, 0]
                            category_complete_status["core_science_ph"] = True
                            category_complete_status["core_experiment_ph"] = True

                            category_complete_status["software"] = True

                            return True

                        # 물리 + 생물 조합
                        if category_complete_status["core_science_ph"] == True:
                            # 화학, 화실을 자유학점에 넣어준다
                            my_classified_courses_credit["others"][1] += \
                                my_classified_courses_credit["core_science_ch"][1] + \
                                my_classified_courses_credit["core_experiment_ch"][1]
                            my_classified_courses["others"].append(
                                my_classified_courses["core_science_ch"])
                            my_classified_courses["others"].append(
                                my_classified_courses["core_experiment_ch"])

                            # 화학, 화실 삭제
                            my_classified_courses_credit["core_science_ch"][1] = 0
                            my_classified_courses_credit["core_experiment_ch"][1] = 0
                            my_classified_courses["core_science_ch"] = []
                            my_classified_courses["core_experiment_ch"] = [
                            ]

                            # 최소, 최대 이수학점 0으로 만들어주고 이수 여부 true로
                            classified_courses_credit["core_science_ch"] = [
                                0, 0]
                            classified_courses_credit["core_experiment_ch"] = [
                                0, 0]
                            category_complete_status["core_science_ch"] = True
                            category_complete_status["core_experiment_ch"] = True

                            category_complete_status["software"] = True

                            return True

        # 컴프 안 듣고 소기코를 들었을 경우
        # 물화생 전부 충족해야 함
        elif my_classified_courses_credit["core_science_cp"][1] == 0 and my_classified_courses_credit["software"][
            1] == 2:

            category_complete_status["software"] = True

            # 화학
            if my_classified_courses_credit["core_science_ch"][1] == 3:
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_science_ch"] = True
                    category_complete_status["core_experiment_ch"] = True

            # 물리
            if my_classified_courses_credit["core_science_ph"][1] == 3:
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_science_ph"] = True
                    category_complete_status["core_experiment_ph"] = True

            # 생물
            if my_classified_courses_credit["core_science_bs"][1] == 3:
                if my_classified_courses_credit["core_experiment_bs"][1] == 1:
                    category_complete_status["core_science_bs"] = True
                    category_complete_status["core_experiment_bs"] = True

            # 세 개 다 충족했다면
            if category_complete_status["core_science_ch"] == True and category_complete_status[
                "core_science_ph"] == True and category_complete_status["core_science_bs"] == True:
                category_complete_status["core_science_cp"] = True

                return True

        # 컴프, 소기코 둘다 아직 안 들은 경우
        # 아마 1학년...?
        # 일단 각각 충족한 것만 True 해주고 넘긴다
        elif my_classified_courses_credit["core_science_cp"][1] == 0 and my_classified_courses_credit["software"][
            1] == 0:

            # 화학
            if my_classified_courses_credit["core_science_ch"][1] == 3:
                category_complete_status["core_science_ch"] = True
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_experiment_ch"] = True

            # 물리
            if my_classified_courses_credit["core_science_ph"][1] == 3:
                category_complete_status["core_science_ph"] = True
                if my_classified_courses_credit["core_experiment_ch"][1] == 1:
                    category_complete_status["core_experiment_ph"] = True

            # 생물
            if my_classified_courses_credit["core_science_bs"][1] == 3:
                category_complete_status["core_science_bs"] = True
                if my_classified_courses_credit["core_experiment_bs"][1] == 1:
                    category_complete_status["core_experiment_bs"] = True

            return True

    # my_classified_courses_credit에 인정된 학점을 넣어주는 함수
    def insert_accepted_credits():
        for category in ["core_english1", "core_english2", "core_writing",
                         "HUS", "PPE", "humanity",
                         "software",
                         "core_math1", "core_math2",
                         "core_science_ch", "core_experiment_ch",
                         "core_science_ph", "core_experiment_ph",
                         "core_science_bs", "core_experiment_bs",
                         "core_science_cp",
                         "freshman_seminar",
                         "major",
                         "research",
                         "uc_core", "uc_optional1", "uc_optional2", "others",
                         "music", "exercise", "colloquium"]:
            try:
                my_classified_courses_credit[category][0] = min(my_classified_courses_credit[category][1],
                                                                classified_courses_credit[category][1])
            except:
                pass

    # 최소 이수 학점을 충족했는지 확인하고 해당 category의 이수 현황을 True로 바꿔주는 함수
    # 기초과학은 제외 => modifying_basic_science() 에서 진행
    def check_clear_category():
        for category in ["core_english1", "core_english2", "core_writing",
                         "HUS", "PPE", "humanity",
                         "software",
                         "core_math1", "core_math2",
                         "freshman_seminar",
                         "major",
                         "research",
                         "uc_core", "uc_optional1", "uc_optional2", "others",
                         "music", "exercise", "colloquium"]:
            try:
                if my_classified_courses_credit[category][0] >= classified_courses_credit[category][0]:
                    category_complete_status[category] = True
            except:
                pass

    # 각 대분류에 대해서 소분류(category)가 충족됐는지 확인해주는 함수
    # 소분류 중에 하나라도 False면 해당 대분류는 False로 바꿔준다
    def check_clear_large_category():
        for large in ["언어의 기초", "교양", "소프트웨어", "기초수학",
                      "신입생 세미나", "전공", "연구",
                      "대학 공통", "자유학점", "예체능", "콜로퀴움",
                      "분류 안 됨", "F or U"]:
            for category in large_category[large]:
                if category == 'nonclassified_courses' or category == 'F or U':
                    continue
                if category_complete_status[category] == False:
                    large_category_complete_status[large] = False
                    break
        if large_category["소프트웨어"] == True:
            for category in large_category["기초과학"]:
                if category_complete_status[category] == False:
                    large_category_complete_status["기초과학"] = False
                    break
        else:
            large_category_complete_status["기초과학"] = False

    # 졸업학점과 인정되지 않은 학점을 더해서 return 해주는 함수
    # sum_credit = [졸업학점, 인정되지 않은 학점]
    def sum_credits():
        sum_credit = [0, 0]

        for category in ["core_english1", "core_english2", "core_writing",
                         "HUS", "PPE", "humanity",
                         "software",
                         "core_math1", "core_math2",
                         "core_science_ch", "core_experiment_ch",
                         "core_science_ph", "core_experiment_ph",
                         "core_science_bs", "core_experiment_bs",
                         "core_science_cp",
                         "freshman_seminar",
                         "major",
                         "research",
                         "uc_core", "uc_optional1", "uc_optional2", "others",
                         "music", "exercise", "colloquium"]:
            sum_credit[0] += my_classified_courses_credit[category][0]

            for category in ["nonclassified_courses", "F or U"]:
                sum_credit[1] += my_classified_courses_credit[category][0]

            return sum_credit

    if request.method == 'POST':
        try:
            ws = openpyxl.load_workbook(filename=request.FILES['document'].file).active
        except:
            messages.error(request, '올바른 형식의 파일을 업로드해주세요!')
            return redirect("pybo:upload_start")

        my_courses = get_my_courses(ws)
        for my_course_index in range(len(my_courses)):

            if not classify_my_course(my_course_index, my_courses):
                my_classified_courses_credit["nonclassified_courses"] += my_courses[my_course_index][1]
                my_classified_courses["nonclassified_courses"].append(my_courses[my_course_index])
        # 준비된 함수 사용
        # 이수 여부
        modifying_basic_science()
        insert_accepted_credits()
        check_clear_category()
        check_clear_large_category()

        total = sum_credits()

        # 대분류 완료 상황
        #large_category_complete_status
        # 소분류 완료 상황
        #ategory_complete_status
        context = {}

        return render(request, 'pybo/result.html', context)

    return render(request, 'pybo/main_content.html')
